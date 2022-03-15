import openpyxl
import requests
import json

def test_post_datadriven_request():
    API_URL = "http://thetestingworldapi.com/api/studentsDetails"

    file = open("C:/Users/pault/PycharmProjects/PycharmAPIAutomationOne/DATA/CreateStudentInfoOne.json", 'r');
    json_request = json.loads(file.read())
    wk = openpyxl.load_workbook("C:/Users/pault/PycharmProjects/PycharmAPIAutomationOne/DATA/DataDriven.xlsx",data_only=True)
    sh=wk['StudentInfo']
    rows=sh.max_row

    for i in range(2,rows+1):
        cell_first_name=sh.cell(row=i,column=1).value
        cell_mid_name = sh.cell(row=i, column=2).value
        cell_last_name = sh.cell(row=i, column=3).value
        cell_dob = sh.cell(row=i, column=4).value
        json_request['first_name'] = cell_first_name
        json_request['middle_name'] = cell_mid_name
        json_request['last_name'] = cell_last_name
        json_request['date_of_birth'] = cell_dob
        responce = requests.post(API_URL, json_request)
        print(responce.text)
        print(responce.status_code)
        assert responce.status_code == 201
