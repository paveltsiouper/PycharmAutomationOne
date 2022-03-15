import openpyxl
import requests
import json
import jsonpath

class Common:

    def __init__(self,FileNamePath, SheetName):
        global wk
        global sh
        wk=openpyxl.load_workbook(FileNamePath,data_only=True)
        sh=wk[SheetName]

    def fetch_row_count(self):
        rows=sh.max_row
        return rows

    def fetch_column_count(self):
        col=sh.max_column
        return col

    def fetch_key_names(self):
        c=sh.max_column
        list_column_names=[]
        for i in range(1, c+1):
            cell=sh.cell(row=1,column=i)
            list_column_names.insert(i-1,cell.value)
        return list_column_names

    def update_request_with_data(self,rowNumber,jsonRequest,keyList):
        """Mapping Excell data against JSON object"""
        c=sh.max_column
        for i in range(1, c+1):
            cell=sh.cell(row=rowNumber,column=i)
            jsonRequest[keyList[i-1]]=cell.value
        return jsonRequest